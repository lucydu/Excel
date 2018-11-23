__author__ = 'zz'
from openpyxl import load_workbook
from common import  pro_path
from common.read_config import ReadConfig
from common.do_mysql import DoMysql

class DoExcel:
    def __init__(self,file_path):
        self.file_path=file_path

    def do_excel(self):
        flag=ReadConfig().read_config(pro_path.conf_path,'TESTCASE','flag')
        case_id_list=ReadConfig().read_config(pro_path.conf_path,'TESTCASE','case_id_list')

        wb=load_workbook(self.file_path)
        sheet=wb['test_data']

         #第一种方式：
        # #从Excel里面获取初始化手机号
        # tel=self.get_tel()#获取的手机号码是整数

         # #获取完毕之后 更新初始化的手机号码
        # self.update_tel(tel+1)
        #第二种方式：直接从数据库里面去获取数据
        #获取数据库当前存在的最大的手机号
        query='select max(mobilephone) from  member'
        tel=int(DoMysql().do_mysql(query)[0])+1#获取第一个元素

        #调用函数 存储手机号码
        self.used_tel(tel)

        test_data=[]#所有测试数据放到一个列表里面
        for i in range(2,sheet.max_row+1):
            sub_data={}#每一行测试数据存在一个字典里面
            sub_data['CaseId']=sheet.cell(i,1).value
            sub_data['Title']=sheet.cell(i,2).value
            sub_data['Method']=sheet.cell(i,3).value
            sub_data['URL']=sheet.cell(i,4).value

             #思考点？如何去做正确的替换
            if sheet.cell(i,5).value.find('${tel}')!=-1:#find函数
                #做替换
                sub_data['Param']=sheet.cell(i,5).value.replace('${tel}',str(tel))
            else:
                #不做替换
                sub_data['Param']=sheet.cell(i,5).value

               #请求参数
            sub_data['ExpectedResult']=sheet.cell(i,6).value
            test_data.append(sub_data)#添加数据到列表里面

        #根据配置文件的开关来决定运行哪些测试用例
        if flag=='on':#意味着执行所有的用例
            final_data=test_data
        else:#执行指定列表里面的数据？
            final_data=[]#用来存储最后要执行的用例
            for i in eval(case_id_list):
                final_data.append(test_data[i-1])#注意这里的写法？
        return final_data#返回最终的数据

    def write_back(self,sheet_name,row,col,new_value):#写回数据
        wb=load_workbook(self.file_path)
        sheet=wb[sheet_name]
        sheet.cell(row,col).value=new_value
        wb.save(self.file_path)

    def used_tel(self,used_tel):#存储已经使用过的电话号码
         wb=load_workbook(self.file_path)
         sheet=wb['used']#存储手机号码的表单
         sheet.cell(sheet.max_row+1,1).value=used_tel
         wb.save(self.file_path)


    def get_tel(self):#获取写在Excel里面的初始化手机号
        wb=load_workbook(self.file_path)
        sheet=wb['tel']
        tel=sheet.cell(1,1).value
        return tel

    # def update_tel(self,new_tel):#更新初始化手机号码函数
    #     wb=load_workbook(self.file_path)
    #     sheet=wb['tel']
    #     sheet.cell(1,1).value=new_tel
    #     wb.save(self.file_path)


if __name__ == '__main__':
    test_data=DoExcel().do_excel('test_cases.xlsx','test_data')
    print("获取到的Excel的测试数据是{0}".format(test_data))